import random
from xml.dom.minidom import parse
from xml.etree.ElementTree import parse as ps
from xml.etree.ElementTree import SubElement, Element, ElementTree
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import numpy as np
import math
from openpyxl import Workbook, load_workbook
from os import listdir, makedirs
from os import path as ospath
from skimage import io, morphology
from skimage.filters import threshold_otsu
from pandas import DataFrame, options, read_excel
from tifffile import TiffFile
from tifffile.tifffile import imagej_description_metadata
options.mode.chained_assignment = None

def get_imlist(path, pic_format='.tif'):
    return [ospath.join(path, f) for f in sorted(listdir(path)) if f.endswith(pic_format)]

def eight_colors():
    # color_list = [(255, 0, 0, 255), (0, 255, 0, 255), (0, 255, 255, 255), (255, 128, 0, 255),
    #               (255, 255, 0, 255), (128, 128, 0, 255), (244, 222, 179, 255), (255, 0, 255, 255)]
    color_list = [(255, 0, 0, 255), (0, 255, 0, 255), (0, 255, 255, 255), (255, 128, 0, 255),
                  (255, 255, 0, 255), (128, 128, 0, 255), (244, 222, 179, 255), (227, 159, 246, 255)]
    return color_list

def can_convert_to_int(string):
    try:
        int(string)
        return True
    except ValueError:
        return False

def get_spots_locations(idx, SpotsInFrame, ImageData, shape):
    SpotsElement = SpotsInFrame[idx]
    Spots = SpotsElement.getElementsByTagName("Spot")
    whether_mitosis = []
    real_X = []
    X = []
    IDs = []
    colors = []
    pixelwidth = float(ImageData[0].getAttribute("pixelwidth"))
    pixelheight = float(ImageData[0].getAttribute("pixelheight"))
    voxeldepth = float(ImageData[0].getAttribute("voxeldepth"))
    for Spot in Spots:
        x = float(Spot.getAttribute("POSITION_X")) / pixelwidth
        y = float(Spot.getAttribute("POSITION_Y")) / pixelheight
        z = float(Spot.getAttribute("POSITION_Z")) / voxeldepth
        id = Spot.getAttribute("ID")
        color = Spot.getAttribute("MANUAL_SPOT_COLOR")
        mitotic = 1 if color == '-65536' else 0
        if can_convert_to_int(color):
            colors.append(int(color))
        else:
            colors.append(int(-100))
        X.append([int(shape[-1]/10)-float(x)/5, int(shape[2]/2)-float(y), int(shape[1]/2) - float(z)])
        whether_mitosis.append(mitotic)
        real_X.append([float(z), float(y), float(x)])
        IDs.append(id)
    return real_X, X, whether_mitosis, IDs, colors

def reset_ids(base_direc, idx, domTree, start_cell_id, time_intervel):
    # For excels, reset ids, for each of 8 lines, ids are in ascending order by y coordinate
    color_xml = [-13382401, -6710785, -10027162, -39271, -16776961, -6711040, -39322, -205]

    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllSpots = Model[0].findall("AllSpots")[0]
    SpotsInFrame = SubElement(AllSpots, "SpotsInFrame")
    SpotsInFrame.tail = '\n'
    SpotsInFrame.attrib = {"frame": f"{idx}"}
    ImageData = get_settings(base_direc)
    pixelwidth = float(ImageData[0].getAttribute("pixelwidth"))
    pixelheight = float(ImageData[0].getAttribute("pixelheight"))
    voxeldepth = float(ImageData[0].getAttribute("voxeldepth"))

    cells_excel, unit_vector = read_excels(idx, base_direc)
    unit_vector = unit_vector.split(" ")
    unit_vector = list(map(float, unit_vector))
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(base_direc, idx))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    data_0.remove_sheet(the_sheet_0)
    data_0.create_sheet('annotations')
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(base_direc, idx))

    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    row0 = ["ID", "Z", "Y", "X", "mitotic", "line",
            "next_spot_1", "next_spot_2",
            f"{str(unit_vector[0]) + ' ' + str(unit_vector[1]) + ' ' + str(unit_vector[2])}"]  # Columns
    for i in range(len(row0)):
        the_sheet_0.cell(row=1, column=i + 1).value = row0[i]

    for i in range(8):
        globals()[f"cells_0_{i}"] = []
    for i in range(len(cells_excel)):
        globals()[f"cells_0_{int(cells_excel[i][-1])}"].append(cells_excel[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        ys = globals()[f"cells_0_{i}"][:, 2]
        ys = np.array([-y for y in ys])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys.argsort()]

    first_element = [globals()[f"cells_0_{i}"] for i in range(8)]
    joined_cells = np.concatenate((first_element[0], first_element[1], first_element[2], first_element[3],
                                   first_element[4], first_element[5], first_element[6], first_element[7]))

    for j, item in enumerate(joined_cells):
        the_sheet_0.cell(row=j + 2, column=1).value = int(start_cell_id)
        the_sheet_0.cell(row=j + 2, column=2).value = float(item[1])
        the_sheet_0.cell(row=j + 2, column=3).value = float(item[2])
        the_sheet_0.cell(row=j + 2, column=4).value = float(item[3])
        the_sheet_0.cell(row=j + 2, column=5).value = int(item[4])
        the_sheet_0.cell(row=j + 2, column=6).value = int(item[5])

        Spot = SubElement(SpotsInFrame, "Spot")
        Spot.tail = '\n'
        if int(item[4]):# mitotic
            Spot.attrib = {"CONTRAST_CH1": "0.14919749142730468", "FRAME": f"{idx}", "ID": f"{start_cell_id}",
                           "MANUAL_SPOT_COLOR": "-65536", "MAX_INTENSITY_CH1": "112.0",
                           "MEAN_INTENSITY_CH1": "90.11392405063289", "MEDIAN_INTENSITY_CH1": "89.0",
                           "MIN_INTENSITY_CH1": "67.0", "POSITION_T": f"{time_intervel * idx}",
                           "POSITION_X": f"{item[3] * pixelwidth}",
                           "POSITION_Y": f"{item[2] * pixelheight}",
                           "POSITION_Z": f"{item[1] * voxeldepth}",
                           "QUALITY": "-1.0", "RADIUS": "2.0719994878893218",
                           "SNR_CH1": "2.2566243856862824", "STD_INTENSITY_CH1": "10.368823267028697",
                           "TOTAL_INTENSITY_CH1": "7119.0", "VISIBILITY": "1", "name": f"ID{start_cell_id}"}
        else:
            Spot.attrib = {"CONTRAST_CH1": "0.14919749142730468", "FRAME": f"{idx}", "ID": f"{start_cell_id}",
                           "MANUAL_SPOT_COLOR": f"{color_xml[int(item[5])]}", "MAX_INTENSITY_CH1": "112.0",
                           "MEAN_INTENSITY_CH1": "90.11392405063289", "MEDIAN_INTENSITY_CH1": "89.0",
                           "MIN_INTENSITY_CH1": "67.0", "POSITION_T": f"{time_intervel * idx}",
                           "POSITION_X": f"{item[3] * pixelwidth}",
                           "POSITION_Y": f"{item[2] * pixelheight}",
                           "POSITION_Z": f"{item[1] * voxeldepth}",
                           "QUALITY": "-1.0", "RADIUS": "2.0719994878893218",
                           "SNR_CH1": "2.2566243856862824", "STD_INTENSITY_CH1": "10.368823267028697",
                           "TOTAL_INTENSITY_CH1": "7119.0", "VISIBILITY": "1", "name": f"ID{start_cell_id}"}

        start_cell_id += 1

    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(base_direc, idx))

    AllSpots.attrib = {"nspots": f"{start_cell_id}"}

    return domTree, start_cell_id


def reload_excel(directory, id, SpotsInFrame, ImageData, shape):
    print(id)
    real_X, _, whether_mitosis, IDs, _ = get_spots_locations(id, SpotsInFrame, ImageData, shape)
    cells_excel, _ = read_excels(id, directory)
    added_IDs = []
    deleted_IDs = []
    IDs = [float(i) for i in IDs]
    for i, temp in enumerate(IDs):
        if temp not in cells_excel[:, 0]:
            added_IDs.append(i)
    for i, temp in enumerate(cells_excel[:, 0]):
        if temp not in IDs:
            deleted_IDs.append(i)
    if added_IDs != []:
        data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
        data_0.active
        the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
        max_row_0 = the_sheet_0.max_row + 1
        for add_id in added_IDs:
            print(IDs[add_id])
            sorted_ranks = find_smallest_distance(real_X[add_id], cells_excel[:, 1:4])
            the_sheet_0.cell(row=max_row_0, column=1).value = IDs[add_id]
            the_sheet_0.cell(row=max_row_0, column=2).value = real_X[add_id][0]
            the_sheet_0.cell(row=max_row_0, column=3).value = real_X[add_id][1]
            the_sheet_0.cell(row=max_row_0, column=4).value = real_X[add_id][2]
            the_sheet_0.cell(row=max_row_0, column=5).value = whether_mitosis[add_id]
            the_sheet_0.cell(row=max_row_0, column=6).value = cells_excel[sorted_ranks[0]][5]
            max_row_0 += 1
        data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, id))
    if deleted_IDs != []:
        data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
        data_0.active
        the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
        for j, ix in enumerate(deleted_IDs):
            print(cells_excel[ix][0])
            the_sheet_0.delete_rows(ix + 2 - j, 1)
        data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, id))

def reload_excel_from_scratch(directory, id, SpotsInFrame, ImageData, shape, new_ranks):
    real_X, _, whether_mitosis, IDs, colors = get_spots_locations(id, SpotsInFrame, ImageData, shape)
    cells_excel, unit_vector = read_excels(id, directory)
    unit_vector = unit_vector.split(" ")
    unit_vector = list(map(float, unit_vector))
    excel_ids = cells_excel[:, 0]
    excel_lines = cells_excel[:, -1]
    print(f'current xlsx is {id}')
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    data_0.remove_sheet(the_sheet_0)
    data_0.create_sheet('annotations')
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, id))

    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    row0 = ["ID", "Z", "Y", "X", "mitotic", "line",
            "next_spot_1", "next_spot_2",
            f"{str(unit_vector[0]) + ' ' + str(unit_vector[1]) + ' '+ str(unit_vector[2])}"]  # Columns
    for i in range(len(row0)):
        the_sheet_0.cell(row=1, column=i + 1).value = row0[i]
    color_xml = [-13382401, -6710785, -10027162, -39271, -16776961, -6711040, -39322, -205]
    for j, item in enumerate(real_X):
        the_sheet_0.cell(row=j + 2, column=1).value = int(IDs[j])
        the_sheet_0.cell(row=j + 2, column=2).value = float(item[0])
        the_sheet_0.cell(row=j + 2, column=3).value = float(item[1])
        the_sheet_0.cell(row=j + 2, column=4).value = float(item[2])
        the_sheet_0.cell(row=j + 2, column=5).value = int(whether_mitosis[j])
        itemindex_list = np.where(excel_ids == int(IDs[j]))
        if np.size(itemindex_list) == 0:
            which_color = colors[j]
            print(which_color)
            color_index = np.where(color_xml == np.array(which_color))
            print(color_index)
            if np.size(color_index) == 0:
                the_sheet_0.cell(row=j + 2, column=6).value = random.randint(0, 7)
            else:
                the_sheet_0.cell(row=j + 2, column=6).value = new_ranks[id][color_index[0][0]]
        else:
            which_color = colors[j]
            print(which_color)
            color_index = np.where(color_xml == np.array(which_color))

            itemindex = itemindex_list[0][0]
            print(itemindex)

            if np.size(color_index) == 0:
                if which_color == -65536:
                    the_sheet_0.cell(row=j + 2, column=6).value = excel_lines[itemindex]
                else:
                    the_sheet_0.cell(row=j + 2, column=6).value = random.randint(0, 7)
            else:
                the_sheet_0.cell(row=j + 2, column=6).value = new_ranks[id][color_index[0][0]]
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, id))
    data_0.close()

def read_excels(id, directory):
    cells_0 = []
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    max_row_0 = the_sheet_0.max_row

    for j in range(2, max_row_0 + 1):
        mi = [the_sheet_0.cell(row=j, column=1).value, the_sheet_0.cell(row=j, column=2).value,
              the_sheet_0.cell(row=j, column=3).value, the_sheet_0.cell(row=j, column=4).value,
              the_sheet_0.cell(row=j, column=5).value, the_sheet_0.cell(row=j, column=6).value]
        cells_0.append(mi)
    return np.array(cells_0), the_sheet_0.cell(row=1, column=9).value

def read_excels_lines(id, directory):
    cells_0 = []
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    max_row_0 = the_sheet_0.max_row
    shape = length_imgs(directory)
    for j in range(2, max_row_0 + 1):
        mi = [the_sheet_0.cell(row=j, column=1).value,
              int(shape[-1]/10) - float(the_sheet_0.cell(row=j, column=4).value)/5,
              int(shape[2]/2) - float(the_sheet_0.cell(row=j, column=3).value),
              int(shape[1]/2) - float(the_sheet_0.cell(row=j, column=2).value),
              the_sheet_0.cell(row=j, column=5).value,
              the_sheet_0.cell(row=j, column=6).value
              ]
        cells_0.append(mi)
    data_0.close()
    return np.array(cells_0),  the_sheet_0.cell(row=1, column=9).value


def points_projection(unit_vector, X):
    unit_vector = np.array(unit_vector)
    distance = np.array([np.sum(unit_vector * x) for x in X])
    new_X = []
    for i, temp in enumerate(X):
        new_X.append(temp-distance[i]*unit_vector)
    new_X = np.array(new_X)
    mask = [True, False, True]
    X_xz = new_X[:, mask]
    return X_xz

def cross_over(parents):
    offsprings = []
    for i in range(len(parents)):
        for j in range(i+1, len(parents)):
            parents_i = parents[i]
            parents_j = parents[j]
            offspring = (np.array(parents_i) + np.array(parents_j))/2
            offspring_hat = offspring / np.linalg.norm(offspring)
            offsprings.append(list(offspring_hat))
    return offsprings


def cart2sph(coords):
    xy = np.sqrt(coords[0] ** 2 + coords[1] ** 2)  # sqrt(x² + y²)
    theta = np.arctan2(coords[1], coords[0])
    phi = np.arctan2(xy, coords[2])
    return theta, phi

def mutation(offsprings_cross_over):
    cross_over_length = len(offsprings_cross_over)
    index = np.random.choice(cross_over_length, int(cross_over_length))
    for j in index:
        x_random = np.random.uniform(-0.2, 0.2)
        y_random = np.random.uniform(-0.2, 0.2)
        z_random = np.random.uniform(-0.2, 0.2)
        mu_spring = np.array(offsprings_cross_over[j]) + np.array([x_random, y_random, z_random])
        mu_spring_hat = mu_spring / np.linalg.norm(mu_spring)
        offsprings_cross_over.append(list(mu_spring_hat))
    return offsprings_cross_over

def fit_func(population, X):
    positive_vs_negative_x = []
    XZs = np.array([points_projection(chromosome, X) for chromosome in population])

    # print(np.shape(XZs))
    distances = []
    for x in range(len(XZs)):
        mean_XZs = np.average(XZs[x], axis=0)
        dist = np.array([math.hypot(temp[0] - mean_XZs[0], temp[1] - mean_XZs[1]) for temp in XZs[x]])
        distances.append(np.std(dist))
    return distances

def random_three_vector():
    x = np.random.uniform(-0.5, 0.5)
    y = np.random.uniform(0.8, 1.2)
    z = np.random.uniform(-0.5, 0.5)
    v = np.array([x, y, z])
    v_hat = v / np.linalg.norm(v)
    return list(v_hat)

def ga_plane(X):
    sol_per_pop = 10000
    new_population = []
    for i in range(sol_per_pop):
        new_chromosome = random_three_vector()
        if new_chromosome not in new_population:
            new_population.append(new_chromosome)
    # new_population = np.array(new_population)
    num_generation = 3
    num_parents_mating = 50

    for generation in range(num_generation):
        fitness = fit_func(new_population, X)
        parents_ids = np.argpartition(fitness, num_parents_mating)
        parents = [new_population[p_id] for p_id in parents_ids[:num_parents_mating]]
        offspring_cross_over = cross_over(parents)
        offspring_mutation = mutation(offspring_cross_over)
        new_population = list(np.zeros((len(parents) + len(offspring_mutation), 3)))
        for j in range(len(parents) + len(offspring_mutation)):
            new_population[j] = parents[j] if j < len(parents) else offspring_mutation[j - len(parents)]
        if generation == num_generation-1:
            unit_vector_final = new_population[np.argmin(fitness)]
            fig_xz_final = points_projection(unit_vector_final, X)
            return unit_vector_final, fig_xz_final

def k_means_8_line(XZ):
    n_clusters = 8
    kmeans = KMeans(n_clusters=n_clusters, random_state=2018)
    kmeans.fit(XZ)
    pre_y = kmeans.predict(XZ)
    return XZ, pre_y

def length_imgs(directory):
    tif = get_imlist(directory, pic_format='.tif')
    if tif == []:
        return [0]
    else:
        img = io.imread(tif[0])
        img_shape = np.shape(img)
        return img_shape

# Write center of cells into .xlsx excel file
def write_excel(idx, centroids, whether_mitosis, line_number, IDs, unit_vector, directory):
    f = Workbook()
    f.active
    sheets = f.sheetnames
    ws = f[sheets[0]]
    ws.title = "annotations"
    row0 = ["ID", "Z", "Y", "X", "mitotic", "line",
            "next_spot_1", "next_spot_2",
            f"{str(unit_vector[0]) + ' ' + str(unit_vector[1]) + ' '+ str(unit_vector[2])}"]  # Columns
    for i in range(len(row0)):
        ws.cell(row=1, column=i + 1).value = row0[i]

    for j, item in enumerate(centroids):
        ws.cell(row=j + 2, column=1).value = int(IDs[j])
        ws.cell(row=j + 2, column=2).value = float(item[0])
        ws.cell(row=j + 2, column=3).value = float(item[1])
        ws.cell(row=j + 2, column=4).value = float(item[2])
        ws.cell(row=j + 2, column=5).value = int(whether_mitosis[j])
        ws.cell(row=j + 2, column=6).value = int(line_number[j])
    makedirs(f'{directory}/tracking', exist_ok=True)
    f.save(f'{directory}/tracking/points_locations_{idx}.xlsx')
    f.close()

def refine_excel(directory, idx, index, color):
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, idx))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    for i in index:
        ix = i.index()
        the_sheet_0.cell(row=ix+2, column=6).value = color
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, idx))

def delete_indexes(directory, idx, index):
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, idx))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    ixs = []
    for i in index:
        ix = i.index()
        ixs.append(ix)
    ixs = sorted(ixs)
    for j, i in enumerate(ixs):
        the_sheet_0.delete_rows(i+2-j, 1)
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, idx))


def get_settings(directory):
    xml = get_imlist(directory, '.xml')
    domTree = parse(xml[0])
    rootNode = domTree.documentElement
    Settings = rootNode.getElementsByTagName("Settings")
    ImageData = Settings[0].getElementsByTagName("ImageData")
    return ImageData

class ga_processing(object):
    def __init__(self, directory):
        self.directory = directory
    def __call__(self, idx):
        self.read_xml(idx)

    def read_xml(self, idx):
        shape = length_imgs(self.directory)
        xml = get_imlist(self.directory, '.xml')
        domTree = parse(xml[0])
        rootNode = domTree.documentElement
        Model = rootNode.getElementsByTagName("Model")
        AllSpots = Model[0].getElementsByTagName("AllSpots")
        SpotsInFrame = AllSpots[0].getElementsByTagName("SpotsInFrame")
        ImageData = get_settings(self.directory)
        real_X, X, whether_mitosis, IDs, _ = get_spots_locations(idx, SpotsInFrame, ImageData, shape)
        unit_vector, XZ = ga_plane(X)
        _, preds_y = k_means_8_line(XZ)
        write_excel(idx, real_X, whether_mitosis, preds_y, IDs, unit_vector, self.directory)

def xz_lines_from_excel(directory, volume):
    data = load_workbook(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, volume))  # 打开Excel文件读取数据
    data.active
    the_sheet = data.get_sheet_by_name("annotations")  # 通过名称获取
    max_row = the_sheet.max_row
    all_cells = []
    for index in range(2, max_row + 1):
        all_cells.append([the_sheet.cell(row=index, column=2).value,
                          the_sheet.cell(row=index, column=3).value,
                          the_sheet.cell(row=index, column=4).value,
                          the_sheet.cell(row=index, column=6).value
                          ])
    all_cells = np.array(all_cells)
    all_cells_sorted = all_cells[all_cells[:, 1].argsort()]
    line_ranks = []
    cells_locations = []
    for y in range(int(len(all_cells_sorted)/2), len(all_cells_sorted)):
        if all_cells_sorted[y][3] not in line_ranks:
            line_ranks.append(all_cells_sorted[y][3])
            cells_locations.append(all_cells_sorted[y][:3])
        if len(line_ranks) == 8:
            break;
    return cells_locations, line_ranks

def line_correspondence(directory, volume, rank):
    data = load_workbook(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, volume))  # 打开Excel文件读取数据
    data.active
    the_sheet = data.get_sheet_by_name("annotations")  # 通过名称获取
    max_row = the_sheet.max_row
    rank = np.array(rank)
    order = [0, 1, 2, 3, 4, 5, 6, 7]
    for index in range(2, max_row + 1):
        the_sheet.cell(row=index, column=6).value = order[np.argwhere(rank == the_sheet.cell(row=index, column=6).value)[0, 0]]
    data.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, volume))

def strat_from_min_angle(list_temp, old_ranks):
    average_points = [0, 0]
    for number, cent in enumerate(list_temp):
        average_points = [average_points[0]+cent[0], average_points[1]+cent[1]]
    average_points = [average_points[0]/len(list_temp), average_points[1]/len(list_temp)]
    polar_coordinates = []
    for number, cent in enumerate(list_temp):
        x = cent[0] - average_points[0]
        y = cent[1] - average_points[1]
        radius = math.sqrt(math.pow(x, 2) + math.pow(y, 2))
        theta = 180 * math.atan2(y, x) / math.pi
        # Displaying polar coordinates
        if theta < 0:
            theta1 = 360+theta
        else:
            theta1 = theta
        polar_coordinates.append([radius, theta, theta1])
    polar_coordinates = np.array(polar_coordinates)
    ranks_change = np.argsort(polar_coordinates[:, 2])
    new_ranks = [old_ranks[ranks_change[i]] for i in range(len(ranks_change))]
    polar_coordinates = np.array(sorted(polar_coordinates, key=lambda x: x[2]))
    abs_theta = list(polar_coordinates[:, 1])
    abs_theta = [abs(a) for a in abs_theta]
    min_index = np.argmin(abs_theta)
    if min_index == 7:
        temp_last = polar_coordinates[7]
        rank_last = new_ranks[7]
        polar_coordinates = list(polar_coordinates)
        new_ranks = list(new_ranks)
        polar_coordinates.pop()
        new_ranks.pop()
        polar_coordinates.insert(0, temp_last)
        new_ranks.insert(0, rank_last)
    polar_coordinates = np.array(polar_coordinates)
    polar_coordinates = polar_coordinates[:, 0:2]
    return polar_coordinates, average_points, new_ranks

def find_smallest_distance(point, list_temp):
    distance = []
    for i in range(len(list_temp)):
        temp = 0
        for j in range(len(list_temp[0])):
            d_j = point[j]-list_temp[i][j]
            temp += float(d_j**2)
        distance.append(temp)
    distance = np.array(distance)
    distance.reshape(-1)
    sorted_ranks = sorted(range(len(distance)), key=lambda k: distance[k])
    return sorted_ranks

def find_smallest_distance_with_distances(point, list_temp):
    distance = []
    for i in range(len(list_temp)):
        temp = 0
        for j in range(len(list_temp[0])):
            d_j = point[j]-list_temp[i][j]
            temp += float(abs(d_j))
        distance.append(temp)
    distance = np.array(distance)
    distance.reshape(-1)
    sorted_ranks = sorted(range(len(distance)), key=lambda k: distance[k])
    if sorted_ranks[1] == 0 and distance[sorted_ranks[1]]-distance[sorted_ranks[0]] < 5 and distance[sorted_ranks[1]] < 6:
        return sorted_ranks[1], distance[sorted_ranks[0]]
    else:
        return sorted_ranks[0], distance[sorted_ranks[0]]

def find_smallest_y(y, list_temp, whether_0=False):
    distance = []
    for i in range(len(list_temp)):
        d_j = y-list_temp[i]
        distance.append(abs(d_j))
    distance = np.array(distance)
    distance.reshape(-1)
    sorted_ranks = sorted(range(len(distance)), key=lambda k: distance[k])
    if sorted_ranks[1] == 0 and distance[sorted_ranks[1]] - distance[sorted_ranks[0]] < 5.0  \
            and distance[sorted_ranks[1]] < 7 and whether_0:
        return sorted_ranks[1], distance[sorted_ranks[0]]
    else:
        return sorted_ranks[0], distance[sorted_ranks[0]]

def find_most_similar_polar(list_prior, list_current, rank_current):
    polar_coordinates, middle_point, rank_first = strat_from_min_angle(list_current, rank_current)
    smallest = find_smallest_distance(polar_coordinates[0], list_prior)[0]
    # print(f'The smallest distance is {smallest}')
    polar_coordinates_ranked = []
    rank_later = []
    if smallest != 0:
        for start in range(8-smallest, 8):
            polar_coordinates_ranked.append(polar_coordinates[start])
            rank_later.append(rank_first[start])
        for start in range(0, 8-smallest):
            polar_coordinates_ranked.append(polar_coordinates[start])
            rank_later.append(rank_first[start])
    else:
        polar_coordinates_ranked = list(polar_coordinates)
        rank_later = rank_first
    return polar_coordinates_ranked, middle_point, rank_later

def get_mitotic_color(directory):
    path = r"{0}\tracking".format(directory)
    df = None
    for file in listdir(path):
        df_tmp = DataFrame(read_excel(ospath.join(path, file), engine='openpyxl',))
        if df is None:
            df = df_tmp
        else:
            df = df.append(df_tmp)

    color_list = [-13382401, -6710785, -10027162, -39271, -16776961, -6711040, -39322, -205]

    xml_file = get_imlist(directory, pic_format='.xml')

    domTree = ps(xml_file[0])
    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllSpots = Model[0].findall("AllSpots")
    SpotsInFrame = AllSpots[0].findall("SpotsInFrame")

    # print(len(SpotsInFrame))
    num = 0
    for idx in range(len(SpotsInFrame)):
        SpotsElement = SpotsInFrame[idx]
        Spots = SpotsElement.findall("Spot")
        # print(len(Spots))
        # print(idx)
        for Spot in Spots:
            id = Spot.attrib["ID"]
            mitotic_color = df.mitotic[df.ID == int(id)]
            if len(mitotic_color.index) != 0:
                mitotic_color = np.array(mitotic_color)[0]
                if mitotic_color:
                    # print(f'{id} is a mitotic')
                    num += 1
                else:
                    colors = df.line[df.ID == int(id)]
                    colors = np.array(colors)[0]
                    color = color_list[colors]
                    Spot.attrib["MANUAL_SPOT_COLOR"] = str(color)
                    # print(id, Spot.attrib["MANUAL_SPOT_COLOR"])
                    num += 1
    domTree.write(xml_file[0])


def link_lines(img_number, directory):
    centroids_for_each_volume = []
    all_ranks = []
    for i in range(img_number):
        X, ranks = xz_lines_from_excel(directory, i)
        # print(ranks)
        X = np.array(X)
        mask = [True, False, True]
        X_xz = X[:, mask]
        centroids_for_each_volume.append(X_xz)
        all_ranks.append(ranks)

    centroids_before = centroids_for_each_volume[0]
    polar_before, middle_point_before, rank_0 = strat_from_min_angle(centroids_before, all_ranks[0])

    new_ranks = []
    new_ranks.append(rank_0)
    for i in range(1, img_number):
        centroids = centroids_for_each_volume[i]
        polar_points, middle_points, ranks = find_most_similar_polar(polar_before, centroids, all_ranks[i])
        new_ranks.append(ranks)
        polar_before = polar_points
    return new_ranks


def distance_between_points(point1, point2):
    if len(point2) != len(point1):
        print('Error, the length of two points is different')
    else:
        temp = 0
        for i in range(len(point1)):
            d_j = point1[i] - point2[i]
            temp += float(d_j ** 2)
        return np.sqrt(temp)

def nethermost_8_cells(cells_0):
    eight_cells = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
    for i in range(len(cells_0)):
        # print(cells_0[i])
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        ys = globals()[f"cells_0_{i}"][:, 2]
        ys = np.array([-y for y in ys])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys.argsort()]
        eight_cells.append(globals()[f"cells_0_{i}"][0][2:4])
    return eight_cells

def nethermost_8_cells_y(cells_0):
    eight_cells = []
    zs = []
    xs = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
    for i in range(len(cells_0)):
        # print(cells_0[i])
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        ys = globals()[f"cells_0_{i}"][:, 2]
        ys = np.array([-y for y in ys])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys.argsort()]
        eight_cells.append(globals()[f"cells_0_{i}"][0][2])
        zs.append(int(globals()[f"cells_0_{i}"][0][1]))
        xs.append(int(globals()[f"cells_0_{i}"][0][3]))
    return eight_cells, zs, xs

# def affine_calculate(moving, fixed, i,q, xyz):
#     moving_s = sitk.GetImageFromArray(moving)
#     moving_sitk = sitk.Cast(moving_s, sitk.sitkFloat32)
#     fixed_s = sitk.GetImageFromArray(fixed)
#     fixed_sitk = sitk.Cast(fixed_s, sitk.sitkFloat32)
#
#     R = sitk.ImageRegistrationMethod()
#     R.SetMetricAsMeanSquares()
#     R.SetOptimizerAsRegularStepGradientDescent(0.1, .001, 200)
#     # R.SetInitialTransform(sitk.Euler2DTransform())
#     R.SetInitialTransform(sitk.AffineTransform(fixed_sitk.GetDimension()))
#     R.SetInterpolator(sitk.sitkLinear)
#     outTx = R.Execute(fixed_sitk, moving_sitk)
#
#     # print(outTx)
#
#     transformed_point = outTx.TransformPoint(xyz)
#     resampler = sitk.ResampleImageFilter()
#     resampler.SetReferenceImage(fixed_sitk)
#     resampler.SetInterpolator(sitk.sitkLinear)
#     resampler.SetDefaultPixelValue(100)
#     resampler.SetTransform(outTx)
#     out = resampler.Execute(moving_sitk)
#
#     # if i == 0:
#     #     plt.imsave(f'./otsu/{i}_{q}.png', sitk.GetArrayFromImage(out))
#     #     plt.imsave(f'./otsu/{i+1}_{q}.png', fixed)
#     # else:
#     #     plt.imsave(f'./otsu/{i + 1}_{q}.png', fixed)
#
#     print(transformed_point)
#     print(xyz)
#     return transformed_point[1]-xyz[1]

def translation_calculate(moving, fixed, i):
    thresh_m = threshold_otsu(moving)
    binary_m = moving > thresh_m+30
    binary_m = morphology.remove_small_objects(np.array(binary_m, dtype=bool), 20)
    locations_moving = np.argwhere(binary_m != 0)
    locations_moving_y = locations_moving[:, 0]

    thresh_f = threshold_otsu(fixed)
    binary_f = fixed > thresh_f+30
    binary_f = morphology.remove_small_objects(np.array(binary_f, dtype=bool), 20)
    locations_fixed = np.argwhere(binary_f != 0)
    locations_fixed_y = locations_fixed[:, 0]
    # if i == 0:
    #     plt.imsave(f'./otsu/{i}.png', binary_m)
    #     plt.imsave(f'./otsu/{i+1}.png', binary_f)
    # else:
    #     plt.imsave(f'./otsu/{i + 1}.png', binary_f)
    print(np.max(locations_moving_y), np.max(locations_fixed_y))

    off_set = np.max(locations_moving_y) - np.max(locations_fixed_y)
    max_y_moving = np.max(locations_moving_y)
    max_y_fixed = np.max(locations_fixed_y)+off_set
    # print(np.sum(np.square(np.subtract(np.roll(binary_f.astype(int), off_set, axis=1)[max_y_fixed-20:max_y_fixed+20, :],
    #                                    binary_m.astype(int)[max_y_moving-20:max_y_moving+20, :]))))
    return off_set

def translation_mask(moving, fixed):
    locations_moving = np.argwhere(moving != 0)
    locations_moving_y = locations_moving[:, 0]
    locations_fixed = np.argwhere(fixed != 0)
    locations_fixed_y = locations_fixed[:, 0]
    # print(np.max(locations_moving_y), np.max(locations_fixed_y))
    off_set = np.max(locations_moving_y) - np.max(locations_fixed_y)
    return off_set

def translation_y(cells_0, cells_1):
    off_set = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
        globals()[f"cells_1_{i}"] = []
    for i in range(len(cells_0)):
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(len(cells_1)):
        globals()[f"cells_1_{int(cells_1[i][-1])}"].append(cells_1[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        globals()[f"cells_1_{i}"] = np.array(globals()[f"cells_1_{i}"])
        ys_0 = globals()[f"cells_0_{i}"][:, 2]
        ys_0 = np.array([-y for y in ys_0])
        ys_1 = globals()[f"cells_1_{i}"][:, 2]
        ys_1 = np.array([-y for y in ys_1])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys_0.argsort()]
        globals()[f"cells_1_{i}"] = globals()[f"cells_1_{i}"][ys_1.argsort()]
        off_set.append(globals()[f"cells_1_{i}"][0][1] - globals()[f"cells_0_{i}"][0][1])
    return off_set

def reset_excels(id, directory):
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    max_row_0 = the_sheet_0.max_row

    for j in range(2, max_row_0 + 1):
        the_sheet_0.cell(row=j, column=7).value = None
        the_sheet_0.cell(row=j, column=8).value = None
    data_0.save(r'{0}/tracking/points_locations_{1}.xlsx'.format(directory, id))
    data_0.close()

def link_spots_by_y(cells_0, cells_1, id, directory, nethermost_0, nethermost_index_0, off_set=[0 for _ in range(8)]):
    data_0 = DataFrame(read_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), engine='openpyxl', ))
    nethermost_1 = []
    nethermost_index_1 = []
    zs = []
    xs = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
        globals()[f"cells_1_{i}"] = []
    for i in range(len(cells_0)):
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(len(cells_1)):
        globals()[f"cells_1_{int(cells_1[i][-1])}"].append(cells_1[i])
    for i in range(8):

        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        globals()[f"cells_1_{i}"] = np.array(globals()[f"cells_1_{i}"])
        ys_0 = globals()[f"cells_0_{i}"][:, 2]
        ys_0 = np.array([-y for y in ys_0])
        ys_1 = globals()[f"cells_1_{i}"][:, 2]
        ys_1 = np.array([-y for y in ys_1])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys_0.argsort()]
        globals()[f"cells_1_{i}"] = globals()[f"cells_1_{i}"][ys_1.argsort()]
        # print(nethermost_0[i])
        # print(globals()[f"cells_1_{i}"][:, 2:4])
        globals()[f"cells_1_{i}"][:, 2] = globals()[f"cells_1_{i}"][:, 2] + off_set[i]
        # if globals()[f"cells_0_{i}"][nethermost_index_0[i]][0] == 61208:
        #     print("yes")
        whether_0 = nethermost_index_0[i] == 0
        which_one, distance = find_smallest_y(nethermost_0[i], globals()[f"cells_1_{i}"][:, 2], whether_0)
        # print(which_one)
        length_0 = len(globals()[f"cells_0_{i}"])
        length_1 = len(globals()[f"cells_1_{i}"])
        # if i == 2 and id == 110:
        #     print("yes")
        # if globals()[f"cells_0_{i}"][nethermost_index_0[i]][0] == 58370 \
        #         or globals()[f"cells_0_{i}"][nethermost_index_0[i]][0] ==58368:
        #     print("yes")
        if distance > 6.0: #threshold: the nethermost cell of line i in frame t-1 does not have corresponding cell in frame t
            temp, _ = find_smallest_y(globals()[f"cells_1_{i}"][which_one][2],
                                      globals()[f"cells_0_{i}"][nethermost_index_0[i]:length_0, 2])
            if temp != 0:
                nethermost_index_0[i] = temp + nethermost_index_0[i]
            else:
                temp1, _ = find_smallest_y(
                    globals()[f"cells_0_{i}"][nethermost_index_0[i]][2],
                    globals()[f"cells_1_{i}"][which_one:length_1, 2])
                which_one = temp1+which_one
        nethermost_index_1.append(which_one)
        nethermost_1.append(globals()[f"cells_1_{i}"][which_one][2]-off_set[i])
        zs.append(int(globals()[f"cells_1_{i}"][which_one][1]))
        xs.append(int(globals()[f"cells_1_{i}"][which_one][3]))
        if length_0-nethermost_index_0[i] >= length_1-which_one:
            end = length_1
        else:
            end = length_0

        start = which_one
        for j in range(nethermost_index_0[i], end):
            if globals()[f"cells_0_{i}"][j][-2]:
                if globals()[f"cells_1_{i}"][start][-2]: #cells not divided yet
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][
                        0]
                    start += 1
                else:
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][
                        0]
                    start += 1
                    if start >= length_1:
                        break
                    else:
                        data_0.next_spot_2[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = \
                        globals()[f"cells_1_{i}"][start][0]
                        start += 1
            else:
                data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                start += 1
            if start >= length_1:
                break
    data_0.to_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), sheet_name='annotations',
                    index=False)
    return nethermost_1, nethermost_index_1, zs, xs


def link_spots_by_xy(cells_0, cells_1, id, directory, nethermost_0, nethermost_index_0, off_set=0):
    data_0 = DataFrame(read_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), engine='openpyxl', ))
    nethermost_1 = []
    nethermost_index_1 = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
        globals()[f"cells_1_{i}"] = []
    for i in range(len(cells_0)):
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(len(cells_1)):
        globals()[f"cells_1_{int(cells_1[i][-1])}"].append(cells_1[i])
    for i in range(8):

        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        globals()[f"cells_1_{i}"] = np.array(globals()[f"cells_1_{i}"])
        ys_0 = globals()[f"cells_0_{i}"][:, 2]
        ys_0 = np.array([-y for y in ys_0])
        ys_1 = globals()[f"cells_1_{i}"][:, 2]
        ys_1 = np.array([-y for y in ys_1])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys_0.argsort()]
        globals()[f"cells_1_{i}"] = globals()[f"cells_1_{i}"][ys_1.argsort()]
        # print(nethermost_0[i])
        # print(globals()[f"cells_1_{i}"][:, 2:4])
        globals()[f"cells_1_{i}"][:, 2] = globals()[f"cells_1_{i}"][:, 2] + off_set
        # if globals()[f"cells_0_{i}"][nethermost_index_0[i]][0] == 486603:
        #     print("yes")
        which_one, distance = find_smallest_distance_with_distances(nethermost_0[i], globals()[f"cells_1_{i}"][:, 2:4])
        # print(which_one)
        length_0 = len(globals()[f"cells_0_{i}"])
        length_1 = len(globals()[f"cells_1_{i}"])


        if distance > 6.0: #threshold: the nethermost cell of line i in frame t-1 does not have corresponding cell in frame t
            # print(globals()[f"cells_0_{i}"][nethermost_index_0[i]:nethermost_index_0[i] + 3, 2:4])
            temp, _ = find_smallest_distance_with_distances(globals()[f"cells_1_{i}"][which_one][2:4],
                                                            globals()[f"cells_0_{i}"][nethermost_index_0[i]:nethermost_index_0[i]+3, 2:4])
            if temp != 0:
                nethermost_index_0[i] = temp + nethermost_index_0[i]
            else:
                temp1, _ = find_smallest_distance_with_distances(
                    globals()[f"cells_0_{i}"][nethermost_index_0[i]][2:4],
                    globals()[f"cells_1_{i}"][which_one:which_one+3, 2:4])
                which_one = temp1+which_one
        nethermost_index_1.append(which_one)
        nethermost_1.append([globals()[f"cells_1_{i}"][which_one][2]-off_set,
                             globals()[f"cells_1_{i}"][which_one][3]])
        if length_0-nethermost_index_0[i] >= length_1-which_one:
            end = length_1
        else:
            end = length_0

        start = which_one
        for j in range(nethermost_index_0[i], end):
            if globals()[f"cells_0_{i}"][j][-2]:
                if globals()[f"cells_1_{i}"][start][-2]: #cells not divided yet
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][
                        0]
                    start += 1
                else:
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][
                        0]
                    start += 1
                    if start >= length_1:
                        break
                    else:
                        data_0.next_spot_2[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = \
                        globals()[f"cells_1_{i}"][start][0]
                        start += 1
            else:
                data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                start += 1
            if start >= length_1:
                break
    data_0.to_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), sheet_name='annotations',
                    index=False)
    return nethermost_1, nethermost_index_1


def link_spots(cells_0, cells_1, id, directory):
    data_0 = DataFrame(read_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), engine='openpyxl', ))
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
        globals()[f"cells_1_{i}"] = []
    for i in range(len(cells_0)):
        globals()[f"cells_0_{int(cells_0[i][-1])}"].append(cells_0[i])
    for i in range(len(cells_1)):
        globals()[f"cells_1_{int(cells_1[i][-1])}"].append(cells_1[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        globals()[f"cells_1_{i}"] = np.array(globals()[f"cells_1_{i}"])
        ys_0 = globals()[f"cells_0_{i}"][:, 2]
        ys_0 = np.array([-y for y in ys_0])
        ys_1 = globals()[f"cells_1_{i}"][:, 2]
        ys_1 = np.array([-y for y in ys_1])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys_0.argsort()]
        globals()[f"cells_1_{i}"] = globals()[f"cells_1_{i}"][ys_1.argsort()]

        length_0 = len(globals()[f"cells_0_{i}"])
        length_1 = len(globals()[f"cells_1_{i}"])

        if length_0 >= length_1:
            end = length_1
        else:
            end = length_0

        start = 0
        for j in range(0, end):
            if globals()[f"cells_0_{i}"][j][-2]:
                if globals()[f"cells_1_{i}"][start][-2]: #cells not divided yet
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                    start += 1
                else:
                    data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                    start += 1
                    if start >= length_1:
                        break
                    else:
                        data_0.next_spot_2[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                        start += 1
            else:
                data_0.next_spot_1[data_0.ID == globals()[f"cells_0_{i}"][j][0]] = globals()[f"cells_1_{i}"][start][0]
                start += 1
            if start >= length_1:
                break
    data_0.to_excel(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id), sheet_name='annotations',
                    index=False)


def track_cells(length, directory):
    all_spots = []
    for i in range(length):
        cells_0, _ = read_excels(i, directory)
        all_spots.append(cells_0)
    return all_spots

def read_excels_tracked_cell(id, directory):
    cells_0 = []
    data_0 = load_workbook(r'{0}\tracking\points_locations_{1}.xlsx'.format(directory, id))  # 打开Excel文件读取数据
    data_0.active
    the_sheet_0 = data_0.get_sheet_by_name("annotations")  # 通过名称获取
    max_row_0 = the_sheet_0.max_row
    for j in range(2, max_row_0 + 1):
        mi = [the_sheet_0.cell(row=j, column=1).value, the_sheet_0.cell(row=j, column=2).value,
              the_sheet_0.cell(row=j, column=3).value, the_sheet_0.cell(row=j, column=4).value,
              the_sheet_0.cell(row=j, column=5).value, the_sheet_0.cell(row=j, column=6).value,
              the_sheet_0.cell(row=j, column=7).value, the_sheet_0.cell(row=j, column=8).value,
              ]
        cells_0.append(mi)
    return np.array(cells_0)


def edge_in_xml(id, cells_source, cell_target, all_cells_target, time_intervel):
    for j in range(len(all_cells_target)):
       if all_cells_target[j][0] == cell_target:
           displacement = distance_between_points([all_cells_target[j][3], all_cells_target[j][2], all_cells_target[j][1]],
                                                  [cells_source[3], cells_source[2], cells_source[1]])
           return [cells_source[0], cell_target, -1.0, id*time_intervel + time_intervel/2,
                   (all_cells_target[j][3] + cells_source[3])/2, (all_cells_target[j][2] + cells_source[2])/2,
                   (all_cells_target[j][1] + cells_source[1])/2, displacement/time_intervel, displacement]

def rank_line_then_y(cells_i):
    ranked_cells = []
    for i in range(8):
        globals()[f"cells_0_{i}"] = []
    for i in range(len(cells_i)):
        globals()[f"cells_0_{cells_i[i][5]}"].append(cells_i[i])
    for i in range(8):
        globals()[f"cells_0_{i}"] = np.array(globals()[f"cells_0_{i}"])
        ys_0 = globals()[f"cells_0_{i}"][:, 2]
        ys_0 = np.array([-y for y in ys_0])
        globals()[f"cells_0_{i}"] = globals()[f"cells_0_{i}"][ys_0.argsort()]
        for j in range(len(ys_0)):
            ranked_cells.append( globals()[f"cells_0_{i}"][j])
    return ranked_cells


def judge_blank_spots(domTree):
    # Remove contents in xml AllSpots
    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllSpots = Model[0].findall("AllSpots")
    Model[0].remove(AllSpots[0])
    AllSpots = SubElement(Model[0], "AllSpots")
    return domTree

def judge_blank_track(domTree):
    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllTracks = Model[0].findall("AllTracks")
    FilteredTracks = Model[0].findall("FilteredTracks")
    Model[0].remove(AllTracks[0])
    Model[0].remove(FilteredTracks[0])
    AllTracks = SubElement(Model[0], "AllTracks")
    FilteredTracks = SubElement(Model[0], "FilteredTracks")
    # all_tracks = AllTracks[0].findall("Track")
    # filtered_tracks = FilteredTracks.findall("TrackID")
    # for at in all_tracks:
    #     all_tracks.remove(at)
    # # if filtered_tracks != 0:
    # for ft in filtered_tracks:
    #     filtered_tracks.remove(ft)
    return domTree


def tracked_cells_to_xml(id, all_cells, tracked_cells, frames, domTree, track_id, time_intervel):
    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllTracks = Model[0].findall("AllTracks")
    FilteredTracks = Model[0].findall("FilteredTracks")
    all_tracks = AllTracks[0]
    filtered_tracks = FilteredTracks[0]
    # print(len(all_tracks.findall("Track")))
    # print(len(filtered_tracks.findall("TrackID")))

    for i, cell in enumerate(all_cells[id]):
        all_cells_next = all_cells[id+1]
        if cell[0] not in tracked_cells:
            track_id = track_id+1
            name = "Track_{0}".format(track_id)
            NUMBER_SPOTS = 0
            NUMBER_SPLITS = 0
            TRACK_START = id * time_intervel
            TRACK_STOP = frames * time_intervel
            TRACK_DURATION = TRACK_STOP - TRACK_START
            TRACK_DISPLACEMENT = 0
            Track_Index = id
            TRACK_X_LOCATION = cell[3]
            TRACK_Y_LOCATION = cell[2]
            TRACK_Z_LOCATION = cell[1]
            Edges = []
            tracked_cells.append(cell[0])
            cell_next = []
            NUMBER_SPOTS += 1
            if cell[-1] != None:
                cell_next.append(cell[-2])
                cell_next.append(cell[-1])
                tracked_cells.append(cell[-2])
                tracked_cells.append(cell[-1])
                edge1 = edge_in_xml(id, cell, cell[-2], all_cells_next, time_intervel)
                edge2 = edge_in_xml(id, cell, cell[-1], all_cells_next, time_intervel)
                Edges.append(edge1)
                Edges.append(edge2)
                TRACK_DISPLACEMENT += edge1[-1]
                TRACK_DISPLACEMENT += edge2[-1]
                NUMBER_SPOTS += 2
                NUMBER_SPLITS += 1
            else:
                if cell[-2] != None:
                    cell_next.append(cell[-2])
                    tracked_cells.append(cell[-2])
                    edge = edge_in_xml(id, cell, cell[-2], all_cells_next, time_intervel)
                    Edges.append(edge)
                    TRACK_DISPLACEMENT += edge[-1]
                    NUMBER_SPOTS += 1
            for j in range(id+2, frames):
                cell_temp = []
                if cell_next != []:
                    for c_next in cell_next:
                        cells_following = all_cells[j]
                        # print(cells_following[:,0])
                        for a_c_next in all_cells_next:
                            if c_next == a_c_next[0]:
                                if a_c_next[-1] != None and a_c_next[-2] != None:
                                    cell_temp.append(a_c_next[-2])
                                    cell_temp.append(a_c_next[-1])
                                    tracked_cells.append(a_c_next[-2])
                                    tracked_cells.append(a_c_next[-1])
                                    edge1 = edge_in_xml(j, a_c_next, a_c_next[-2], cells_following, time_intervel)
                                    edge2 = edge_in_xml(j, a_c_next, a_c_next[-1], cells_following, time_intervel)
                                    TRACK_DISPLACEMENT += edge1[-1]
                                    TRACK_DISPLACEMENT += edge2[-1]
                                    Edges.append(edge1)
                                    Edges.append(edge2)
                                    NUMBER_SPOTS += 2
                                    NUMBER_SPLITS += 1
                                elif a_c_next[-2] != None and a_c_next[-1] == None:
                                    cell_temp.append(a_c_next[-2])
                                    tracked_cells.append(a_c_next[-2])
                                    edge = edge_in_xml(j, a_c_next, a_c_next[-2], cells_following, time_intervel)
                                    TRACK_DISPLACEMENT += edge[-1]
                                    Edges.append(edge)
                                    NUMBER_SPOTS += 1
                                else:
                                    continue;
                else:
                    TRACK_STOP = j * time_intervel
                    TRACK_DURATION = TRACK_STOP - TRACK_START
                    break;
                cell_next = cell_temp
                all_cells_next = cells_following
            if Edges!=[]:
                # print(Edges)
                track = SubElement(all_tracks, "Track")
                track.tail = '\n'
                track.attrib = {"name": f"{name}", "TRACK_ID": f"{track_id}", "NUMBER_SPOTS": f"{NUMBER_SPOTS}",
                                "NUMBER_GAPS": "0", "LONGEST_GAP": "0", "NUMBER_SPLITS": f"{NUMBER_SPLITS}",
                                "NUMBER_MERGES": "0", "NUMBER_COMPLEX": "0",
                                "TRACK_DURATIO": f"{float(TRACK_DURATION)}","TRACK_START": f"{float(TRACK_START)}",
                                "TRACK_STOP": f"{float(TRACK_STOP)}", "TRACK_DISPLACEMENT": f"{TRACK_DISPLACEMENT}",
                                "TRACK_INDEX": f"{Track_Index}", "TRACK_X_LOCATION": f"{float(TRACK_X_LOCATION)}",
                                "TRACK_Y_LOCATION": f"{TRACK_Y_LOCATION}",
                                "TRACK_Z_LOCATION": f"{float(TRACK_Z_LOCATION)}",
                                "TRACK_MEAN_SPEED": "0.012972888129428225", "TRACK_MAX_SPEED": "0.012972888129428225",
                                "TRACK_MIN_SPEED": "0.012972888129428225", "TRACK_MEDIAN_SPEED": "0.012972888129428225",
                                "TRACK_STD_SPEED": "NaN", "TRACK_MEAN_QUALITY":"-1.0", "TRACK_MAX_QUALITY": "-1.0",
                                "TRACK_MIN_QUALITY": "-1.0", "TRACK_MEDIAN_QUALITY": "-1.0", "TRACK_STD_QUALITY": "0.0"}
                for eg in Edges:
                    eg_attributes = {"SPOT_SOURCE_ID": f"{eg[0]}", "SPOT_TARGET_ID": f"{eg[1]}",
                                     "LINK_COST": f"{eg[2]}",  "EDGE_TIME": f"{eg[3]}", "EDGE_X_LOCATION": f"{eg[4]}",
                                     "EDGE_Y_LOCATION": f"{eg[5]}", "EDGE_Z_LOCATION": f"{eg[6]}",
                                     "VELOCITY": f"{eg[7]}", "DISPLACEMENT": f"{eg[8]}"}
                    edg = SubElement(track, "Edge")
                    edg.attrib = eg_attributes
                    edg.tail = '\n'
            if Edges!=[]:
                filtered_tracks_attrib = {"TRACK_ID": f"{track_id}"}
                ft = SubElement(filtered_tracks, "TrackID", attrib=filtered_tracks_attrib)
                ft.tail = '\n'

    return tracked_cells, track_id, domTree

def create_blank_xml(directory, img_shape):
    tif_image = get_imlist(directory)[0]
    folder_name = ospath.dirname(tif_image)
    filename = ospath.basename(tif_image)
    if not ospath.exists(r'{0}/{1}.xml'.format(directory, filename[0])):
        tif = TiffFile(tif_image)
        tags = tif.pages[0].tags
        ij_description = tags['ImageDescription'].value
        x_resolution = tags['XResolution'].value
        y_resolution = tags['YResolution'].value
        ij_description_metadata = imagej_description_metadata(ij_description)
        Trackmate = Element('Trackmate')
        Trackmate.tail = '\n'
        Trackmate.attrib = {"version": "7.6.0"}
        Log = SubElement(Trackmate, "Log")
        model_attrib = {"spatialunits": "micron", "timeunits": "sec"}
        Model = SubElement(Trackmate, "Model")
        Model.tail = '\n'
        Model.attrib = model_attrib
        Settings = SubElement(Trackmate, "Settings")
        Settings.tail = '\n'
        ImageData = SubElement(Settings, "ImageData")
        ImageData.attrib = {"filename":f"{filename}", "folder":f"{folder_name}", "height":f"{img_shape[2]}",
                            "nframes":f"{img_shape[0]}", "nslices":f"{img_shape[1]}",
                            "pixelheight":f"{y_resolution[1]/y_resolution[0]}",
                            "pixelwidth":f"{x_resolution[1]/x_resolution[0]}",
                            "timeinterval":f"{ij_description_metadata['finterval']}",
                            "voxeldepth":f"{ij_description_metadata['spacing']}",
                            "width":f"{img_shape[3]}"}
        BasicSettings = SubElement(Settings, "BasicSettings")
        BasicSettings.attrib = {"tend":f"{img_shape[0]-1}", "tstart":"0",
                                "xend": f"{img_shape[3]-1}", "xstart": "0",
                                "yend": f"{img_shape[2]-1}", "ystart": "0",
                                "zend":f"{img_shape[1]-1}", "zstart": "0"}
        InitialSpotFilter = SubElement(Settings, "InitialSpotFilter")
        InitialSpotFilter.attrib = {"feature": "QUALITY", "isabove": "true", "value": "0.0"}
        SpotFilterCollection = SubElement(Settings, "SpotFilterCollection")
        TrackerSettings = SubElement(Settings, "TrackerSettings")
        TrackerSettings.attrib = {"TRACKER_NAME": "MANUAL_TRACKER"}
        TrackFilterCollection = SubElement(Settings, "TrackFilterCollection")
        AnalyzerCollection = SubElement(Settings, "AnalyzerCollection")
        AnalyzerCollection.tail = '\n'
        SpotAnalyzers = SubElement(AnalyzerCollection, "SpotAnalyzers")
        SpotAnalyzers.tail = '\n'
        Analyzer1 = SubElement(SpotAnalyzers, "Analyzer")
        Analyzer1.attrib = {"key":"Manual spot color"}
        Analyzer2 = SubElement(SpotAnalyzers, "Analyzer")
        Analyzer2.attrib = {"key": "Spot intensity"}
        Analyzer3 = SubElement(SpotAnalyzers, "Analyzer")
        Analyzer3.attrib = {"key": "Spot contrast and SNR"}
        EdgeAnalyzers = SubElement(AnalyzerCollection, "EdgeAnalyzers")
        EdgeAnalyzers.tail = '\n'
        Analyzer_edg = SubElement(EdgeAnalyzers, "Analyzer")
        Analyzer_edg.attrib = { "key":"Directional change"}
        Analyzer_edg1 = SubElement(EdgeAnalyzers, "Analyzer")
        Analyzer_edg1.attrib = { "key":"Edge speed"}
        Analyzer_edg2 = SubElement(EdgeAnalyzers, "Analyzer")
        Analyzer_edg2.attrib = {"key": "Edge target"}
        Analyzer_edg3 = SubElement(EdgeAnalyzers, "Analyzer")
        Analyzer_edg3.attrib = {"key": "Edge location"}
        Analyzer_edg4 = SubElement(EdgeAnalyzers, "Analyzer")
        Analyzer_edg4.attrib = {"key": "Manual edge color"}
        TrackAnalyzers = SubElement(AnalyzerCollection, "TrackAnalyzers")
        TrackAnalyzers.tail = '\n'
        Analyzer_track = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track.attrib = {"key": "Branching analyzer"}
        Analyzer_track1 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track1.attrib = {"key": "Track duration"}
        Analyzer_track2 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track2.attrib = {"key": "Track index"}
        Analyzer_track3 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track3.attrib = {"key": "Track location"}
        Analyzer_track4 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track4.attrib = {"key": "Track speed"}
        Analyzer_track5 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track5.attrib = {"key": "Track quality"}
        Analyzer_track6 = SubElement(TrackAnalyzers, "Analyzer")
        Analyzer_track6.attrib = {"key": "Track motility analysis"}
        FeatureDeclarations = SubElement(Model, "FeatureDeclarations")
        FeatureDeclarations.tail = '\n'
        SpotFeatures = SubElement(FeatureDeclarations, "SpotFeatures")
        SpotFeatures.tail = '\n'
        Feature = SubElement(SpotFeatures, "Feature")
        Feature.attrib = {"dimension": "QUALITY", "feature": "QUALITY", "isint": "false",
                          "name": "Quality", "shortname": "Quality" }
        Feature1 = SubElement(SpotFeatures, "Feature")
        Feature1.attrib = { "dimension": "POSITION", "feature": "POSITION_X", "isint": "false",
                            "name": "X", "shortname": "X"}
        Feature2 = SubElement(SpotFeatures, "Feature")
        Feature2.attrib = { "dimension": "POSITION", "feature": "POSITION_Y", "isint": "false",
                            "name": "Y", "shortname": "Y"}
        Feature3 = SubElement(SpotFeatures, "Feature")
        Feature3.attrib = { "dimension": "POSITION", "feature": "POSITION_Z", "isint": "false",
                            "name": "Z", "shortname": "Z"}
        Feature4 = SubElement(SpotFeatures, "Feature")
        Feature4.attrib = {"dimension": "TIME", "feature": "POSITION_T", "isint": "false",
                           "name": "T", "shortname": "T"}
        Feature5 = SubElement(SpotFeatures, "Feature")
        Feature5.attrib = {"dimension": "NONE", "feature": "FRAME", "isint": "true",
                           "name": "Frame", "shortname": "Frame"}
        Feature6 = SubElement(SpotFeatures, "Feature")
        Feature6.attrib = {"dimension": "LENGTH", "feature": "RADIUS", "isint": "false",
                           "name": "Radius", "shortname": "R"}
        Feature7 = SubElement(SpotFeatures, "Feature")
        Feature7.attrib = {"dimension": "NONE", "feature": "VISIBILITY", "isint": "true",
                           "name": "Visibility", "shortname": "Visibility"}
        Feature8 = SubElement(SpotFeatures, "Feature")
        Feature8.attrib = {"dimension": "NONE", "feature": "MANUAL_SPOT_COLOR", "isint": "true",
                           "name": "Manual spot color", "shortname": "Spot color"}
        Feature9 = SubElement(SpotFeatures, "Feature")
        Feature9.attrib = {"dimension": "INTENSITY", "feature": "MEAN_INTENSITY_CH1", "isint": "false",
                           "name": "Mean intensity ch1", "shortname": "Mean ch1"}
        Feature10 = SubElement(SpotFeatures, "Feature")
        Feature10.attrib = {"dimension": "INTENSITY", "feature": "MEDIAN_INTENSITY_CH1", "isint": "false",
                           "name": "Median intensity ch1", "shortname": "Median ch1"}
        Feature11 = SubElement(SpotFeatures, "Feature")
        Feature11.attrib = {"dimension": "INTENSITY", "feature": "MIN_INTENSITY_CH1", "isint": "false",
                            "name": "Min intensity ch1", "shortname": "Min ch1"}
        Feature12 = SubElement(SpotFeatures, "Feature")
        Feature12.attrib = {"dimension": "INTENSITY", "feature": "MAX_INTENSITY_CH1", "isint": "false",
                            "name": "Max intensity ch1", "shortname": "Max ch1"}
        Feature13 = SubElement(SpotFeatures, "Feature")
        Feature13.attrib = {"dimension": "INTENSITY", "feature": "TOTAL_INTENSITY_CH1", "isint": "false",
                            "name": "Sum intensity ch1", "shortname": "Sum ch1"}
        Feature14 = SubElement(SpotFeatures, "Feature")
        Feature14.attrib = {"dimension": "INTENSITY", "feature": "STD_INTENSITY_CH1", "isint": "false",
                            "name": "Std intensity ch1", "shortname": "Std ch1"}
        Feature15 = SubElement(SpotFeatures, "Feature")
        Feature15.attrib = {"dimension": "NONE", "feature": "CONTRAST_CH1", "isint": "false",
                            "name": "Contrast ch1", "shortname": "Ctrst ch1"}
        Feature16 = SubElement(SpotFeatures, "Feature")
        Feature16.attrib = {"dimension": "NONE", "feature": "SNR_CH1", "isint": "false",
                            "name": "Signal/Noise ratio ch1", "shortname": "SNR ch1"}
        EdgeFeatures = SubElement(FeatureDeclarations, "EdgeFeatures")
        EdgeFeatures.tail = '\n'
        Feature_edg = SubElement(EdgeFeatures, "Feature")
        Feature_edg.attrib = {"dimension": "NONE", "feature": "SPOT_SOURCE_ID", "isint": "true",
                          "name": "Source spot ID", "shortname": "Source ID"}
        Feature_edg1 = SubElement(EdgeFeatures, "Feature")
        Feature_edg1.attrib = {"dimension": "NONE", "feature": "SPOT_TARGET_ID", "isint": "true",
                          "name": "Target spot ID", "shortname": "Target ID"}
        Feature_edg2 = SubElement(EdgeFeatures, "Feature")
        Feature_edg2.attrib = {"dimension": "COST", "feature": "LINK_COST", "isint": "false",
                               "name": "Edge cost", "shortname": "Cost"}
        Feature_edg3 = SubElement(EdgeFeatures, "Feature")
        Feature_edg3.attrib = {"dimension": "ANGLE_RATE", "feature": "DIRECTIONAL_CHANGE_RATE", "isint": "false",
                               "name": "Directional change rate", "shortname": "&#947; rate"}
        Feature_edg4 = SubElement(EdgeFeatures, "Feature")
        Feature_edg4.attrib = {"dimension": "VELOCITY", "feature": "SPEED", "isint": "false",
                               "name": "Speed", "shortname": "Speed"}
        Feature_edg5 = SubElement(EdgeFeatures, "Feature")
        Feature_edg5.attrib = {"dimension": "LENGTH", "feature": "DISPLACEMENT", "isint": "false",
                               "name": "Displacement", "shortname": "Disp."}
        Feature_edg6 = SubElement(EdgeFeatures, "Feature")
        Feature_edg6.attrib = {"dimension": "TIME", "feature": "EDGE_TIME", "isint": "false",
                               "name": "Edge time", "shortname": "Edge T"}
        Feature_edg7 = SubElement(EdgeFeatures, "Feature")
        Feature_edg7.attrib = {"dimension": "POSITION", "feature": "EDGE_X_LOCATION", "isint": "false",
                               "name": "Edge X", "shortname": "Edge X"}
        Feature_edg8 = SubElement(EdgeFeatures, "Feature")
        Feature_edg8.attrib = {"dimension": "POSITION", "feature": "EDGE_Y_LOCATION", "isint": "false",
                               "name": "Edge Y", "shortname": "Edge Y"}
        Feature_edg9 = SubElement(EdgeFeatures, "Feature")
        Feature_edg9.attrib = {"dimension": "POSITION", "feature": "EDGE_Z_LOCATION", "isint": "false",
                               "name": "Edge Z", "shortname": "Edge Z"}
        Feature_edg10 = SubElement(EdgeFeatures, "Feature")
        Feature_edg10.attrib = {"dimension": "NONE", "feature": "MANUAL_EGE_COLOR", "isint": "true",
                               "name": "Manual edge color", "shortname": "Edge color"}
        Feature_edg11 = SubElement(EdgeFeatures, "Feature")
        Feature_edg11.attrib = {"dimension": "NONE", "feature": "MANUAL_EDGE_COLOR", "isint": "true",
                                "name": "Manual edge color", "shortname": "Edge color"}

        TrackFeatures = SubElement(FeatureDeclarations, "TrackFeatures")
        TrackFeatures.tail = '\n'
        Feature_track = SubElement(TrackFeatures, "Feature")
        Feature_track.attrib = {"dimension": "NONE", "feature": "TRACK_INDEX", "isint": "true",
                              "name": "Track index", "shortname": "Index"}
        Feature_track1 = SubElement(TrackFeatures, "Feature")
        Feature_track1.attrib = {"dimension": "NONE", "feature": "TRACK_ID", "isint": "true",
                              "name": "Track ID", "shortname": "ID"}
        Feature_track2 = SubElement(TrackFeatures, "Feature")
        Feature_track2.attrib = {"dimension": "TIME", "feature": "DIVISION_TIME_MEAN", "isint": "false",
                                 "name": "Mean cell division time", "shortname": "Mean div. time"}
        Feature_track3 = SubElement(TrackFeatures, "Feature")
        Feature_track3.attrib = {"dimension": "TIME", "feature": "DIVISION_TIME_STD", "isint": "false",
                                 "name": "Std cell division time", "shortname": "Std div. time"}
        Feature_track4 = SubElement(TrackFeatures, "Feature")
        Feature_track4.attrib = {"dimension": "NONE", "feature": "NUMBER_SPOTS", "isint": "true",
                                 "name": "Number of spots in track", "shortname": "N spots"}
        Feature_track5 = SubElement(TrackFeatures, "Feature")
        Feature_track5.attrib = {"dimension": "NONE", "feature": "NUMBER_GAPS", "isint": "true",
                                 "name": "Number of gaps", "shortname": "N gaps"}
        Feature_track6 = SubElement(TrackFeatures, "Feature")
        Feature_track6.attrib = {"dimension": "NONE", "feature": "NUMBER_SPLITS", "isint": "true",
                                 "name": "Number of split events", "shortname": "N splits"}
        Feature_track7 = SubElement(TrackFeatures, "Feature")
        Feature_track7.attrib = {"dimension": "NONE", "feature": "NUMBER_MERGES", "isint": "true",
                                 "name": "Number of merge events", "shortname": "N merges"}
        Feature_track8 = SubElement(TrackFeatures, "Feature")
        Feature_track8.attrib = {"dimension": "NONE", "feature": "NUMBER_COMPLEX", "isint": "true",
                                 "name": "Number of complex points", "shortname": "N complex"}
        Feature_track9 = SubElement(TrackFeatures, "Feature")
        Feature_track9.attrib = {"dimension": "NONE", "feature": "LONGEST_GAP", "isint": "true",
                                 "name": "Longest gap", "shortname": "Lgst gap"}
        Feature_track10 = SubElement(TrackFeatures, "Feature")
        Feature_track10.attrib = {"dimension": "TIME", "feature": "TRACK_DURATION", "isint": "false",
                                 "name": "Track duration", "shortname": "Duration"}
        Feature_track11 = SubElement(TrackFeatures, "Feature")
        Feature_track11.attrib = {"dimension": "TIME", "feature": "TRACK_START", "isint": "false",
                                  "name": "Track start", "shortname": "Track start"}
        Feature_track12 = SubElement(TrackFeatures, "Feature")
        Feature_track12.attrib = {"dimension": "TIME", "feature": "TRACK_STOP", "isint": "false",
                                  "name": "Track stop", "shortname": "Track stop"}
        Feature_track13 = SubElement(TrackFeatures, "Feature")
        Feature_track13.attrib = {"dimension": "LENGTH", "feature": "TRACK_DISPLACEMENT", "isint": "false",
                                  "name": "Track displacement", "shortname": "Track disp."}
        Feature_track14 = SubElement(TrackFeatures, "Feature")
        Feature_track14.attrib = {"dimension": "POSITION", "feature": "TRACK_X_LOCATION", "isint": "false",
                                  "name": "Track mean X", "shortname": "Track X"}
        Feature_track15 = SubElement(TrackFeatures, "Feature")
        Feature_track15.attrib = {"dimension": "POSITION", "feature": "TRACK_Y_LOCATION", "isint": "false",
                                  "name": "Track mean Y", "shortname": "Track Y"}
        Feature_track16 = SubElement(TrackFeatures, "Feature")
        Feature_track16.attrib = {"dimension": "POSITION", "feature": "TRACK_Z_LOCATION", "isint": "false",
                                  "name": "Track mean Z", "shortname": "Track Z"}
        Feature_track17 = SubElement(TrackFeatures, "Feature")
        Feature_track17.attrib = {"dimension": "VELOCITY", "feature": "TRACK_MEAN_SPEED", "isint": "false",
                                  "name": "Track mean speed", "shortname": "Mean sp."}
        Feature_track18 = SubElement(TrackFeatures, "Feature")
        Feature_track18.attrib = {"dimension": "VELOCITY", "feature": "TRACK_MAX_SPEED", "isint": "false",
                                  "name": "Track max speed", "shortname": "Max speed"}
        Feature_track19 = SubElement(TrackFeatures, "Feature")
        Feature_track19.attrib = {"dimension": "VELOCITY", "feature": "TRACK_MIN_SPEED", "isint": "false",
                                  "name": "Track min speed", "shortname": "Min speed"}
        Feature_track20 = SubElement(TrackFeatures, "Feature")
        Feature_track20.attrib = {"dimension": "VELOCITY", "feature": "TRACK_MEDIAN_SPEED", "isint": "false",
                                  "name": "Track median speed", "shortname": "Med. speed"}
        Feature_track21 = SubElement(TrackFeatures, "Feature")
        Feature_track21.attrib = {"dimension": "VELOCITY", "feature": "TRACK_STD_SPEED", "isint": "false",
                                  "name": "Track std speed", "shortname": "Std speed"}
        Feature_track22 = SubElement(TrackFeatures, "Feature")
        Feature_track22.attrib = {"dimension": "QUALITY", "feature": "TRACK_MEAN_QUALITY", "isint": "false",
                                  "name": "Track mean quality", "shortname": "Mean Q"}
        Feature_track23 = SubElement(TrackFeatures, "Feature")
        Feature_track23.attrib = {"dimension": "LENGTH", "feature": "TOTAL_DISTANCE_TRAVELED", "isint": "false",
                                  "name": "Total distance traveled", "shortname": "Total dist."}
        Feature_track24 = SubElement(TrackFeatures, "Feature")
        Feature_track24.attrib = {"dimension": "LENGTH", "feature": "MAX_DISTANCE_TRAVELED", "isint": "false",
                                  "name": "Max distance traveled", "shortname": "Max dist."}
        Feature_track25 = SubElement(TrackFeatures, "Feature")
        Feature_track25.attrib = {"dimension": "NONE", "feature": "CONFINMENT_RATIO", "isint": "false",
                                  "name": "Confinment ratio", "shortname": "Cfn. ratio"}
        Feature_track26 = SubElement(TrackFeatures, "Feature")
        Feature_track26.attrib = {"dimension": "VELOCITY", "feature": "MEAN_STRAIGHT_LINE_SPEED", "isint": "false",
                                  "name": "Mean straight line speed", "shortname": "Mn. v. line"}
        Feature_track27 = SubElement(TrackFeatures, "Feature")
        Feature_track27.attrib = {"dimension": "NONE", "feature": "LINEARITY_OF_FORWARD_PROGRESSION", "isint": "false",
                                  "name": "Linearity of forward progression", "shortname": "Fwd. progr."}
        Feature_track28 = SubElement(TrackFeatures, "Feature")
        Feature_track28.attrib = {"dimension": "ANGLE_RATE", "feature": "MEAN_DIRECTIONAL_CHANGE_RATE", "isint": "false",
                                  "name": "Mean directional change rate", "shortname": "Mn. &#947; rate"}
        Feature_track29 = SubElement(TrackFeatures, "Feature")
        Feature_track29.attrib = {"dimension": "NONE", "feature": "CONFINEMENT_RATIO", "isint": "false",
                                  "name": "Confinement ratio", "shortname": "Cfn. ratio"}
        AllSpots = SubElement(Model, "AllSpots")
        AllTracks = SubElement(Model, "AllTracks")
        FilteredTracks = SubElement(Model, "FilteredTracks")

        GUIState = SubElement(Trackmate, "GUIState")
        GUIState.attrib = {"state": "ConfigureViews"}
        DisplaySettings = SubElement(Trackmate, "DisplaySettings")
        DisplaySettings.tail = '\n'
        DisplaySettings.text = "{ \n " \
                               "\"name\": \"CurrentDisplaySettings\", " \
                               "\n \"spotUniformColor\": \"0, 255, 255, 255\", " \
                               "\n \"spotColorByType\": \"SPOTS\",\n \"spotColorByFeature\": \"MANUAL_SPOT_COLOR\"," \
                               "\n \"spotDisplayRadius\": 1.0," \
                               "\n \"spotDisplayedAsRoi\": true," \
                               "\n \"spotMin\": 0.0," \
                               "\n \"spotMax\": 10.0," \
                               "\n \"spotShowName\": false," \
                               "\n \"trackMin\": 0.0," \
                               "\n \"trackMax\": 10.0," \
                               "\n \"trackColorByType\": \"TRACKS\"," \
                               "\n \"trackColorByFeature\": \"TRACK_INDEX\"," \
                               "\n \"trackUniformColor\": \"204, 204, 51, 255\"," \
                               "\n \"undefinedValueColor\": \"0, 0, 0, 255\"," \
                               "\n \"missingValueColor\": \"89, 89, 89, 255\"," \
                               "\n \"highlightColor\": \"51, 230, 51, 255\"," \
                               "\n \"trackDisplayMode\": \"FULL\"," \
                               "\n \"colormap\": \"Jet\"," \
                               "\n \"limitZDrawingDepth\": true," \
                               "\n \"drawingZDepth\": 10.0," \
                               "\n \"fadeTracks\": true," \
                               "\n \"fadeTrackRange\": 30," \
                               "\n \"useAntialiasing\": true," \
                               "\n \"spotVisible\": true," \
                               "\n \"trackVisible\": true," \
                               "\n \"font\": {" \
                               "\n \"name\": \"Arial\"," \
                               "\n \"style\": 1," \
                               "\n \"size\": 12," \
                               "\n \"pointSize\": 12.0," \
                               "\n \"fontSerializedDataVersion\": 1" \
                               "\n }," \
                               "\n \"lineThickness\": 1.0," \
                               "\n \"selectionLineThickness\": 4.0," \
                               "\n \"trackschemeBackgroundColor1\": \"128, 128, 128, 255\"," \
                               "\n \"trackschemeBackgroundColor2\": \"192, 192, 192, 255\"," \
                               "\n \"trackschemeForegroundColor\": \"0, 0, 0, 255\"," \
                               "\n \"trackschemeDecorationColor\": \"0, 0, 0, 255\"," \
                               "\n \"trackschemeFillBox\": false," \
                               "\n \"spotFilled\": false," \
                               "\n \"spotTransparencyAlpha\": 1.0" \
                               "\n }"

        tree = ElementTree(Trackmate)
        tree.write(f"{directory}/{ospath.splitext(filename)[0]}.xml")

def xml_from_centroids(idx, common_centroids, mitotic_centroids, domTree, cell_id, time_intervel, directory):
    rootNode = domTree.getroot()
    Model = rootNode.findall("Model")
    AllSpots = Model[0].findall("AllSpots")[0]

    SpotsInFrame = SubElement(AllSpots, "SpotsInFrame")
    SpotsInFrame.tail = '\n'
    SpotsInFrame.attrib = {"frame": f"{idx}"}
    ImageData = get_settings(directory)
    pixelwidth = float(ImageData[0].getAttribute("pixelwidth"))
    pixelheight = float(ImageData[0].getAttribute("pixelheight"))
    voxeldepth = float(ImageData[0].getAttribute("voxeldepth"))
    print(len(common_centroids))
    print(len(mitotic_centroids))
    for i in range(len(common_centroids)):
        Spot = SubElement(SpotsInFrame, "Spot")
        Spot.tail = '\n'
        Spot.attrib = {"CONTRAST_CH1": "0.14919749142730468", "FRAME": f"{idx}", "ID": f"{cell_id}",
                       "MANUAL_SPOT_COLOR": "-13382401", "MAX_INTENSITY_CH1": "112.0",
                       "MEAN_INTENSITY_CH1": "90.11392405063289", "MEDIAN_INTENSITY_CH1":"89.0",
                       "MIN_INTENSITY_CH1": "67.0", "POSITION_T": f"{time_intervel*idx}",
                       "POSITION_X": f"{common_centroids[i][-1] * pixelwidth}",
                       "POSITION_Y": f"{common_centroids[i][1] * pixelheight}",
                       "POSITION_Z": f"{common_centroids[i][0] * voxeldepth}",
                       "QUALITY": "-1.0", "RADIUS": "2.0719994878893218",
                       "SNR_CH1": "2.2566243856862824", "STD_INTENSITY_CH1": "10.368823267028697",
                       "TOTAL_INTENSITY_CH1": "7119.0", "VISIBILITY": "1", "name": f"ID{cell_id}"}
        cell_id += 1
    for j in range(len(mitotic_centroids)):
        Spot = SubElement(SpotsInFrame, "Spot")
        Spot.tail = '\n'
        Spot.attrib = {"CONTRAST_CH1": "0.14919749142730468", "FRAME": f"{idx}", "ID": f"{cell_id}",
                       "MANUAL_SPOT_COLOR": "-65536", "MAX_INTENSITY_CH1": "112.0",
                       "MEAN_INTENSITY_CH1": "90.11392405063289", "MEDIAN_INTENSITY_CH1": "89.0",
                       "MIN_INTENSITY_CH1": "67.0", "POSITION_T": f"{time_intervel * idx}",
                       "POSITION_X": f"{mitotic_centroids[j][-1] * pixelwidth}",
                       "POSITION_Y": f"{mitotic_centroids[j][1] * pixelheight}",
                       "POSITION_Z": f"{mitotic_centroids[j][0] * voxeldepth}",
                       "QUALITY": "-1.0", "RADIUS": "2.0719994878893218",
                       "SNR_CH1": "2.2566243856862824", "STD_INTENSITY_CH1": "10.368823267028697",
                       "TOTAL_INTENSITY_CH1": "7119.0", "VISIBILITY": "1", "name": f"ID{cell_id}"}
        cell_id += 1
    AllSpots.attrib = {"nspots": f"{cell_id+1}"}
    return cell_id, domTree

